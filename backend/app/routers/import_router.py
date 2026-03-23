"""
Import router — CSV and Excel file import.
Prefix /import, tags ["import"].
"""

import asyncio
import csv
import io
from datetime import date
from decimal import Decimal, InvalidOperation
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, UploadFile, status
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.deps import get_current_user
from app.database import get_db
from app.models.item import PLATFORM_VALUES, Item
from app.models.user import User

router = APIRouter(prefix="/import", tags=["import"])

DbDep = Annotated[AsyncSession, Depends(get_db)]
CurrentUserDep = Annotated[User, Depends(get_current_user)]

# ---------------------------------------------------------------------------
# Date parsing helper
# ---------------------------------------------------------------------------

_DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y")


def _parse_date(value: str) -> date | None:
    """Try multiple date formats, return None if none match."""
    value = value.strip()
    if not value:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return date.fromisoformat(value) if fmt == "%Y-%m-%d" else _strptime_date(value, fmt)
        except ValueError:
            continue
    return None


def _strptime_date(value: str, fmt: str) -> date:
    from datetime import datetime as _dt
    return _dt.strptime(value, fmt).date()


def _parse_decimal(value: str) -> Decimal | None:
    """Parse a decimal string, return None if empty or invalid."""
    value = value.strip().replace(",", ".")
    if not value:
        return None
    try:
        return Decimal(value)
    except InvalidOperation:
        return None


# ---------------------------------------------------------------------------
# Row → Item creation helper
# ---------------------------------------------------------------------------

_REQUIRED_COLS = ("nom", "plateforme", "prix_achat", "date_achat")

_PLATFORM_MAP = {p: p for p in PLATFORM_VALUES}


def _normalise_platform(raw: str) -> str:
    raw = raw.strip().lower()
    return _PLATFORM_MAP.get(raw, "autres")


async def _create_item_from_row(
    db: AsyncSession,
    user_id,
    row: dict,
) -> None:
    """Create a single Item from a parsed row dict. Raises ValueError on bad data."""
    name = row.get("nom", "").strip()
    if not name:
        raise ValueError("Colonne 'nom' manquante ou vide.")

    platform_raw = row.get("plateforme", "").strip()
    platform = _normalise_platform(platform_raw)

    purchase_price_raw = row.get("prix_achat", "").strip()
    purchase_price = _parse_decimal(purchase_price_raw)
    if purchase_price is None:
        raise ValueError(f"'prix_achat' invalide : {purchase_price_raw!r}")

    purchase_date_raw = row.get("date_achat", "").strip()
    purchase_date = _parse_date(purchase_date_raw)
    if purchase_date is None:
        raise ValueError(f"'date_achat' invalide : {purchase_date_raw!r}")

    sale_price = _parse_decimal(row.get("prix_vente", ""))
    sale_date = _parse_date(row.get("date_vente", ""))
    platform_fees = _parse_decimal(row.get("frais_plateforme", "")) or Decimal("0.00")
    shipping_cost = _parse_decimal(row.get("frais_port", "")) or Decimal("0.00")
    description = row.get("description", "").strip() or None

    item_status = "sold" if sale_price is not None else "unsold"

    item = Item(
        user_id=user_id,
        name=name,
        description=description,
        platform=platform,
        status=item_status,
        purchase_price=purchase_price,
        purchase_date=purchase_date,
        sale_price=sale_price,
        sale_date=sale_date,
        platform_fees=platform_fees,
        shipping_cost=shipping_cost,
    )
    db.add(item)


# ---------------------------------------------------------------------------
# Sync CSV parse (for asyncio.to_thread)
# ---------------------------------------------------------------------------


def _parse_csv_bytes(content: bytes) -> list[dict]:
    """Parse CSV bytes into a list of row dicts."""
    # Try utf-8-sig first (Excel BOM), then utf-8, then latin-1
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Encodage du fichier CSV non reconnu.")

    reader = csv.DictReader(io.StringIO(text))
    # Normalise headers: strip + lowercase
    rows = []
    for raw_row in reader:
        rows.append({k.strip().lower(): v for k, v in raw_row.items()})
    return rows


# ---------------------------------------------------------------------------
# Sync Excel parse (for asyncio.to_thread)
# ---------------------------------------------------------------------------


def _parse_excel_bytes(content: bytes) -> list[dict]:
    """Parse Excel bytes into a list of row dicts (same schema as CSV)."""
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)

    # Prefer "Import Revendu" sheet, fall back to first
    if "Import Revendu" in wb.sheetnames:
        ws = wb["Import Revendu"]
    else:
        ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        return []

    # Normalise headers
    headers = [str(h).strip().lower() if h is not None else "" for h in header_row]

    results = []
    for row in rows_iter:
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue  # skip blank rows
        row_dict = {}
        for key, cell in zip(headers, row):
            value = "" if cell is None else str(cell).strip()
            row_dict[key] = value
        results.append(row_dict)

    return results


# ---------------------------------------------------------------------------
# POST /import/csv
# ---------------------------------------------------------------------------


@router.post("/csv", summary="Importer des articles depuis un fichier CSV")
async def import_csv(
    file: UploadFile,
    db: DbDep,
    current_user: CurrentUserDep,
) -> dict:
    if file.content_type not in (
        "text/csv",
        "text/plain",
        "application/csv",
        "application/octet-stream",
    ) and not (file.filename or "").endswith(".csv"):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Le fichier doit être un CSV (.csv).",
        )

    content = await file.read()
    if not content:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Le fichier est vide.",
        )

    try:
        rows = await asyncio.to_thread(_parse_csv_bytes, content)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=str(exc),
        ) from exc

    imported = 0
    errors: list[dict] = []

    for row_num, row in enumerate(rows, start=2):  # row 1 = header
        try:
            await _create_item_from_row(db, current_user.id, row)
            imported += 1
        except Exception as exc:  # noqa: BLE001
            errors.append({"row": row_num, "error": str(exc)})

    if imported > 0:
        await db.commit()

    return {"imported": imported, "errors": errors}


# ---------------------------------------------------------------------------
# POST /import/excel
# ---------------------------------------------------------------------------


@router.post("/excel", summary="Importer des articles depuis un fichier Excel (.xlsx)")
async def import_excel(
    file: UploadFile,
    db: DbDep,
    current_user: CurrentUserDep,
) -> dict:
    filename = file.filename or ""
    if not filename.endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Le fichier doit être un Excel (.xlsx).",
        )

    content = await file.read()
    if not content:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Le fichier est vide.",
        )

    try:
        rows = await asyncio.to_thread(_parse_excel_bytes, content)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Impossible de lire le fichier Excel : {exc}",
        ) from exc

    imported = 0
    errors: list[dict] = []

    for row_num, row in enumerate(rows, start=2):
        try:
            await _create_item_from_row(db, current_user.id, row)
            imported += 1
        except Exception as exc:  # noqa: BLE001
            errors.append({"row": row_num, "error": str(exc)})

    if imported > 0:
        await db.commit()

    return {"imported": imported, "errors": errors}
