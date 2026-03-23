"""
Export service — generates Excel and PDF reports for the user.
All heavy I/O-bound (sync) operations are wrapped with asyncio.to_thread().
"""

import io
from datetime import date, datetime, timezone
from decimal import Decimal

from app.models.item import Item
from app.models.user import User

# DAC7 thresholds (same constants as item_service)
DAC7_TRANSACTIONS: int = 30
DAC7_RECEIPTS: Decimal = Decimal("2000.00")

# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

_INDIGO = "6366F1"
_LIGHT_GRAY = "F8F9FA"
_WHITE = "FFFFFF"
_YELLOW_PALE = "FFFACD"
_GREEN = "22C55E"
_ORANGE = "F97316"
_RED = "EF4444"


def _dac7_status(rx_pct: float) -> tuple[str, str]:
    """Return (label, colour-hex) for the DAC7 status."""
    if rx_pct >= 100.0:
        return "DÉPASSÉ", _RED
    if rx_pct >= 85.0:
        return "DANGER", _RED
    if rx_pct >= 70.0:
        return "ATTENTION", _ORANGE
    return "OK", _GREEN


def _fmt_eur(value: Decimal | None) -> str:
    if value is None:
        return ""
    return f"{value:,.2f} €".replace(",", " ")


def _fmt_date(d: date | None) -> str:
    if d is None:
        return ""
    return d.strftime("%d/%m/%Y")


# ---------------------------------------------------------------------------
# Excel export (openpyxl)
# ---------------------------------------------------------------------------


def _build_excel_report(
    user: User,
    items: list[Item],
    year: int,
) -> bytes:
    """Synchronous function — call via asyncio.to_thread()."""
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        GradientFill,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ------------------------------------------------------------------
    # Shared style helpers
    # ------------------------------------------------------------------
    indigo_fill = PatternFill(fill_type="solid", fgColor=_INDIGO)
    gray_fill = PatternFill(fill_type="solid", fgColor=_LIGHT_GRAY)
    thin_side = Side(style="thin", color="CCCCCC")
    thin_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )

    def header_font() -> Font:
        return Font(name="Calibri", bold=True, color=_WHITE, size=11)

    def body_font(bold: bool = False) -> Font:
        return Font(name="Calibri", bold=bold, size=10)

    def center_align() -> Alignment:
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left_align() -> Alignment:
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ------------------------------------------------------------------
    # Sheet 1 — Résumé
    # ------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Résumé"
    ws1.sheet_view.showGridLines = False

    # Logo / titre
    ws1.merge_cells("B2:F2")
    logo_cell = ws1["B2"]
    logo_cell.value = "REVENDU"
    logo_cell.font = Font(name="Calibri", bold=True, size=28, color=_INDIGO)
    logo_cell.alignment = left_align()
    ws1.row_dimensions[2].height = 36

    ws1.merge_cells("B3:F3")
    sub_cell = ws1["B3"]
    sub_cell.value = f"Rapport fiscal — Année {year}"
    sub_cell.font = Font(name="Calibri", size=13, color="555555")
    sub_cell.alignment = left_align()
    ws1.row_dimensions[3].height = 22

    ws1.merge_cells("B4:F4")
    user_cell = ws1["B4"]
    user_cell.value = f"Utilisateur : {user.full_name}"
    user_cell.font = Font(name="Calibri", size=11, italic=True, color="777777")
    user_cell.alignment = left_align()
    ws1.row_dimensions[4].height = 18

    ws1.merge_cells("B5:F5")
    date_cell = ws1["B5"]
    date_cell.value = f"Généré le : {datetime.now(timezone.utc).strftime('%d/%m/%Y à %H:%M')} UTC"
    date_cell.font = Font(name="Calibri", size=9, color="999999")
    date_cell.alignment = left_align()
    ws1.row_dimensions[5].height = 14

    # Stats
    sold_items = [i for i in items if i.status == "sold"]
    total_sold = len(sold_items)
    gross_receipts = sum((i.sale_price or Decimal("0")) for i in sold_items)
    net_profit = sum((i.net_profit or Decimal("0")) for i in sold_items)
    rx_pct = min(float(gross_receipts) / float(DAC7_RECEIPTS) * 100.0, 200.0)
    tx_pct = min(total_sold / DAC7_TRANSACTIONS * 100.0, 200.0)
    status_label, _ = _dac7_status(max(rx_pct, tx_pct))

    stats_data = [
        ("Total ventes", str(total_sold)),
        ("Recettes brutes", _fmt_eur(gross_receipts)),
        ("Bénéfice net", _fmt_eur(net_profit)),
        (f"Seuil DAC7 (recettes)", f"{rx_pct:.1f}%"),
        (f"Seuil DAC7 (transactions)", f"{tx_pct:.1f}%"),
        ("Statut DAC7", status_label),
    ]

    ws1.row_dimensions[7].height = 18
    for row_idx, (label, value) in enumerate(stats_data, start=8):
        lbl = ws1.cell(row=row_idx, column=2, value=label)
        lbl.font = Font(name="Calibri", bold=True, size=10, color="333333")
        lbl.alignment = left_align()
        lbl.border = thin_border
        lbl.fill = gray_fill if row_idx % 2 == 0 else PatternFill(fill_type="solid", fgColor=_WHITE)

        val = ws1.cell(row=row_idx, column=3, value=value)
        val.font = body_font()
        val.alignment = left_align()
        val.border = thin_border
        val.fill = gray_fill if row_idx % 2 == 0 else PatternFill(fill_type="solid", fgColor=_WHITE)

        ws1.row_dimensions[row_idx].height = 20

    ws1.column_dimensions["A"].width = 3
    ws1.column_dimensions["B"].width = 32
    ws1.column_dimensions["C"].width = 22
    ws1.column_dimensions["D"].width = 5

    # ------------------------------------------------------------------
    # Sheet 2 — Ventes détaillées
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("Ventes détaillées")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A2"

    headers = [
        "Nom",
        "Plateforme",
        "Statut",
        "Date achat",
        "Prix achat",
        "Date vente",
        "Prix vente",
        "Frais plateforme",
        "Frais port",
        "Recette brute",
        "Bénéfice net",
    ]

    euro_fmt = '# ##0.00 "€"'
    col_widths = [30, 14, 10, 13, 14, 13, 14, 16, 12, 14, 14]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = header_font()
        cell.fill = indigo_fill
        cell.alignment = center_align()
        cell.border = thin_border
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    ws2.row_dimensions[1].height = 22

    for row_idx, item in enumerate(items, start=2):
        row_fill = PatternFill(fill_type="solid", fgColor=_WHITE if row_idx % 2 == 0 else _LIGHT_GRAY)
        row_data = [
            item.name,
            item.platform,
            item.status,
            _fmt_date(item.purchase_date),
            float(item.purchase_price) if item.purchase_price else None,
            _fmt_date(item.sale_date),
            float(item.sale_price) if item.sale_price else None,
            float(item.platform_fees) if item.platform_fees else 0.0,
            float(item.shipping_cost) if item.shipping_cost else 0.0,
            float(item.gross_receipt) if item.gross_receipt else None,
            float(item.net_profit) if item.net_profit else None,
        ]
        money_cols = {5, 7, 8, 9, 10, 11}  # 1-based column index

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font()
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = left_align()
            if col_idx in money_cols and isinstance(value, float):
                cell.number_format = euro_fmt

        ws2.row_dimensions[row_idx].height = 18

    # ------------------------------------------------------------------
    # Sheet 3 — Déclaration fiscale
    # ------------------------------------------------------------------
    ws3 = wb.create_sheet("Déclaration fiscale")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 3
    ws3.column_dimensions["B"].width = 20
    ws3.column_dimensions["C"].width = 18
    ws3.column_dimensions["D"].width = 18
    ws3.column_dimensions["E"].width = 18

    # Title
    ws3.merge_cells("B2:E2")
    t = ws3["B2"]
    t.value = f"Récapitulatif par plateforme — {year}"
    t.font = Font(name="Calibri", bold=True, size=14, color=_INDIGO)
    t.alignment = left_align()
    ws3.row_dimensions[2].height = 28

    plat_headers = ["Plateforme", "Nb transactions", "Recettes brutes", "Bénéfice net"]
    for col_idx, h in enumerate(plat_headers, start=2):
        cell = ws3.cell(row=4, column=col_idx, value=h)
        cell.font = header_font()
        cell.fill = indigo_fill
        cell.alignment = center_align()
        cell.border = thin_border
    ws3.row_dimensions[4].height = 22

    platform_map: dict[str, dict] = {}
    for item in sold_items:
        p = item.platform
        if p not in platform_map:
            platform_map[p] = {"count": 0, "gross": Decimal("0"), "profit": Decimal("0")}
        platform_map[p]["count"] += 1
        platform_map[p]["gross"] += item.sale_price or Decimal("0")
        platform_map[p]["profit"] += item.net_profit or Decimal("0")

    for row_idx, (platform, data) in enumerate(
        sorted(platform_map.items(), key=lambda x: x[1]["gross"], reverse=True),
        start=5,
    ):
        row_fill = PatternFill(fill_type="solid", fgColor=_WHITE if row_idx % 2 == 0 else _LIGHT_GRAY)
        row_data = [platform, data["count"], float(data["gross"]), float(data["profit"])]
        for col_idx, value in enumerate(row_data, start=2):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font()
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = left_align()
            if col_idx in {4, 5}:
                cell.number_format = '# ##0.00 "€"'
        ws3.row_dimensions[row_idx].height = 18

    next_row = 5 + len(platform_map) + 2

    # Declaration section
    ws3.merge_cells(f"B{next_row}:E{next_row}")
    t2 = ws3[f"B{next_row}"]
    t2.value = "Pour votre déclaration 2042"
    t2.font = Font(name="Calibri", bold=True, size=12, color=_INDIGO)
    t2.alignment = left_align()
    ws3.row_dimensions[next_row].height = 24

    decl_data = [
        ("Total recettes brutes à déclarer", _fmt_eur(gross_receipts)),
        ("Total bénéfice net", _fmt_eur(net_profit)),
        ("Nombre de transactions", str(total_sold)),
        ("Seuil DAC7 recettes", f"{rx_pct:.1f}% (seuil : 2 000 €)"),
        ("Seuil DAC7 transactions", f"{tx_pct:.1f}% (seuil : 30)"),
    ]
    for i, (label, value) in enumerate(decl_data, start=next_row + 1):
        row_fill = PatternFill(fill_type="solid", fgColor=_WHITE if i % 2 == 0 else _LIGHT_GRAY)
        lbl = ws3.cell(row=i, column=2, value=label)
        lbl.font = Font(name="Calibri", bold=True, size=10)
        lbl.fill = row_fill
        lbl.border = thin_border
        lbl.alignment = left_align()
        val = ws3.cell(row=i, column=3, value=value)
        val.font = body_font()
        val.fill = row_fill
        val.border = thin_border
        val.alignment = left_align()
        ws3.row_dimensions[i].height = 18

    footer_row = next_row + 1 + len(decl_data) + 1
    ws3.merge_cells(f"B{footer_row}:E{footer_row}")
    footer = ws3[f"B{footer_row}"]
    footer.value = "Préparé par Revendu - revendu.fr"
    footer.font = Font(name="Calibri", size=9, italic=True, color="999999")
    footer.alignment = left_align()

    # ------------------------------------------------------------------
    # Serialize
    # ------------------------------------------------------------------
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Excel template (openpyxl)
# ---------------------------------------------------------------------------


def _build_excel_template() -> bytes:
    """Synchronous function — call via asyncio.to_thread()."""
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "Import Revendu"
    ws.sheet_view.showGridLines = False

    indigo_fill = PatternFill(fill_type="solid", fgColor="6366F1")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFACD")
    thin_side = Side(style="thin", color="CCCCCC")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    headers = [
        ("Nom*", True, 25, "Nom de l'article (obligatoire). Ex: Nike Air Max 90"),
        ("Plateforme*", True, 16, "Plateforme de vente (obligatoire). Valeurs: vinted, leboncoin, ebay, vestiaire, autres"),
        ("Prix achat*", True, 14, "Prix d'achat en euros (obligatoire). Ex: 45.00"),
        ("Date achat*", True, 14, "Date d'achat au format YYYY-MM-DD, DD/MM/YYYY ou DD-MM-YYYY (obligatoire)"),
        ("Prix vente", False, 14, "Prix de vente en euros. Laisser vide si non vendu. Ex: 72.00"),
        ("Date vente", False, 14, "Date de vente. Même formats que Date achat. Laisser vide si non vendu."),
        ("Frais plateforme", False, 18, "Frais prélevés par la plateforme en euros. Ex: 3.50"),
        ("Frais port", False, 12, "Frais d'envoi en euros. Ex: 0.00"),
        ("Description", False, 30, "Description libre (optionnel). Ex: Taille 42, très bon état"),
    ]

    # Validation liste plateforme (colonne B = col 2)
    dv = DataValidation(
        type="list",
        formula1='"vinted,leboncoin,ebay,vestiaire,autres"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Plateforme invalide",
        error="Choisissez: vinted, leboncoin, ebay, vestiaire, autres",
    )
    ws.add_data_validation(dv)
    dv.sqref = "B3:B10000"

    for col_idx, (label, mandatory, width, comment_text) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        cell.fill = indigo_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 24

        # Yellow background on mandatory columns
        if mandatory:
            cell.fill = PatternFill(fill_type="solid", fgColor="4F46E5")  # darker indigo for mandatory

        # Comment
        comment = Comment(comment_text, "Revendu")
        comment.width = 250
        comment.height = 60
        cell.comment = comment

    # Example row
    example_row = [
        "Nike Air Max",
        "vinted",
        45.00,
        "2024-01-15",
        72.00,
        "2024-02-01",
        3.50,
        0.00,
        "Taille 42",
    ]
    money_cols_template = {3, 5, 7, 8}
    for col_idx, value in enumerate(example_row, start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.font = Font(name="Calibri", size=10, italic=True, color="555555")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if col_idx in money_cols_template and isinstance(value, float):
            cell.number_format = '# ##0.00 "€"'
    ws.row_dimensions[2].height = 18

    # ------------------------------------------------------------------
    # Sheet 2 — Mode d'emploi
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("Mode d'emploi")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 3
    ws2.column_dimensions["B"].width = 25
    ws2.column_dimensions["C"].width = 50

    ws2.merge_cells("B2:C2")
    t = ws2["B2"]
    t.value = "Mode d'emploi — Import Revendu"
    t.font = Font(name="Calibri", bold=True, size=16, color="6366F1")
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[2].height = 30

    instructions = [
        ("", ""),
        ("Étape 1", "Remplissez la feuille \"Import Revendu\" avec vos articles."),
        ("Étape 2", "Les colonnes marquées * sont obligatoires (en-tête violet foncé)."),
        ("Étape 3", "La colonne Plateforme propose une liste déroulante : sélectionnez parmi vinted, leboncoin, ebay, vestiaire, autres."),
        ("Étape 4", "Formats de dates acceptés : YYYY-MM-DD, DD/MM/YYYY ou DD-MM-YYYY."),
        ("Étape 5", "Les prix sont en euros (ex: 45.00). N'incluez pas le symbole €."),
        ("Étape 6", "Laissez vide Prix vente et Date vente pour les articles non vendus."),
        ("Étape 7", "Sauvegardez le fichier au format .xlsx et importez-le sur revendu.fr"),
        ("", ""),
        ("Support", "En cas de problème : support@revendu.fr"),
    ]
    for row_idx, (key, value) in enumerate(instructions, start=3):
        ws2.cell(row=row_idx, column=2, value=key).font = Font(name="Calibri", bold=True, size=10, color="333333")
        ws2.cell(row=row_idx, column=3, value=value).font = Font(name="Calibri", size=10)
        ws2.row_dimensions[row_idx].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# PDF export (reportlab)
# ---------------------------------------------------------------------------


def _build_pdf_report(
    user: User,
    items: list[Item],
    year: int,
) -> bytes:
    """Synchronous function — call via asyncio.to_thread()."""
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (
        HRFlowable,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = io.BytesIO()

    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=1.5 * cm,
        bottomMargin=2.5 * cm,
        title=f"Revendu — Rapport fiscal {year}",
        author="Revendu - revendu.fr",
    )

    page_width = A4[0] - 4 * cm  # usable width

    indigo = colors.HexColor("#6366F1")
    light_gray = colors.HexColor("#F8F9FA")
    dark_gray = colors.HexColor("#374151")
    green = colors.HexColor("#22C55E")
    orange = colors.HexColor("#F97316")
    red = colors.HexColor("#EF4444")
    white = colors.white

    styles = getSampleStyleSheet()

    style_normal = ParagraphStyle(
        "NormalFR",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        textColor=dark_gray,
        leading=13,
    )
    style_small = ParagraphStyle(
        "SmallFR",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        textColor=colors.HexColor("#9CA3AF"),
        leading=11,
    )
    style_section = ParagraphStyle(
        "SectionFR",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=12,
        textColor=indigo,
        leading=18,
        spaceAfter=4,
    )
    style_note = ParagraphStyle(
        "NoteFR",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8.5,
        textColor=dark_gray,
        leading=13,
    )

    # ------------------------------------------------------------------
    # Data aggregation
    # ------------------------------------------------------------------
    sold_items = [i for i in items if i.status == "sold"]
    total_sold = len(sold_items)
    gross_receipts = sum((i.sale_price or Decimal("0")) for i in sold_items)
    net_profit_total = sum((i.net_profit or Decimal("0")) for i in sold_items)
    rx_pct = min(float(gross_receipts) / float(DAC7_RECEIPTS) * 100.0, 200.0)
    tx_pct = min(total_sold / DAC7_TRANSACTIONS * 100.0, 200.0)
    max_pct = max(rx_pct, tx_pct)
    dac7_label, _ = _dac7_status(max_pct)

    story = []

    # ------------------------------------------------------------------
    # Header banner
    # ------------------------------------------------------------------
    gen_date = datetime.now(timezone.utc).strftime("%d/%m/%Y à %H:%M UTC")
    header_data = [
        [
            Paragraph(
                f'<font name="Helvetica-Bold" size="24" color="white">REVENDU</font>',
                ParagraphStyle("h1", alignment=TA_LEFT, leading=28),
            ),
            Paragraph(
                f'<font name="Helvetica-Bold" size="14" color="white">Rapport fiscal {year}</font><br/>'
                f'<font name="Helvetica" size="9" color="#FFFFFF">{user.full_name} — Généré le {gen_date}</font>',
                ParagraphStyle("h2", alignment=TA_RIGHT, leading=18),
            ),
        ]
    ]
    header_table = Table(header_data, colWidths=[page_width * 0.45, page_width * 0.55])
    header_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), indigo),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (0, -1), 14),
                ("RIGHTPADDING", (-1, 0), (-1, -1), 14),
                ("TOPPADDING", (0, 0), (-1, -1), 12),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                ("ROUNDEDCORNERS", [4, 4, 4, 4]),
            ]
        )
    )
    story.append(header_table)
    story.append(Spacer(1, 0.5 * cm))

    # ------------------------------------------------------------------
    # KPIs — 4 boxes
    # ------------------------------------------------------------------
    story.append(Paragraph("Résumé exécutif", style_section))

    kpi_col = page_width / 4

    def _kpi_color(label: str) -> colors.Color:
        if label == "DÉPASSÉ" or label == "DANGER":
            return red
        if label == "ATTENTION":
            return orange
        return green

    kpi_data = [
        [
            Paragraph(
                f'<font name="Helvetica-Bold" size="18" color="#6366F1">{total_sold}</font><br/>'
                f'<font name="Helvetica" size="8" color="#6B7280">Total ventes</font>',
                ParagraphStyle("kpi", alignment=TA_CENTER, leading=22),
            ),
            Paragraph(
                f'<font name="Helvetica-Bold" size="14" color="#6366F1">{_fmt_eur(gross_receipts)}</font><br/>'
                f'<font name="Helvetica" size="8" color="#6B7280">Recettes brutes</font>',
                ParagraphStyle("kpi", alignment=TA_CENTER, leading=22),
            ),
            Paragraph(
                f'<font name="Helvetica-Bold" size="14" color="#6366F1">{_fmt_eur(net_profit_total)}</font><br/>'
                f'<font name="Helvetica" size="8" color="#6B7280">Bénéfice net</font>',
                ParagraphStyle("kpi", alignment=TA_CENTER, leading=22),
            ),
            Paragraph(
                f'<font name="Helvetica-Bold" size="14" color="#{_kpi_color(dac7_label).hexval()[1:]}">{dac7_label}</font><br/>'  # noqa: E501
                f'<font name="Helvetica" size="8" color="#6B7280">Statut DAC7</font>',
                ParagraphStyle("kpi", alignment=TA_CENTER, leading=22),
            ),
        ]
    ]
    kpi_table = Table(kpi_data, colWidths=[kpi_col] * 4)
    kpi_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), light_gray),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
            ]
        )
    )
    story.append(kpi_table)
    story.append(Spacer(1, 0.4 * cm))

    # ------------------------------------------------------------------
    # Détail des ventes
    # ------------------------------------------------------------------
    story.append(Paragraph("Détail des ventes", style_section))

    sale_headers = [
        "Nom",
        "Plateforme",
        "Date vente",
        "Prix achat",
        "Prix vente",
        "Frais",
        "Bénéfice net",
    ]
    col_widths_pdf = [
        page_width * 0.22,
        page_width * 0.12,
        page_width * 0.12,
        page_width * 0.13,
        page_width * 0.13,
        page_width * 0.13,
        page_width * 0.15,
    ]

    header_style = ParagraphStyle("th", fontName="Helvetica-Bold", fontSize=8, textColor=white, alignment=TA_CENTER, leading=11)
    cell_style = ParagraphStyle("td", fontName="Helvetica", fontSize=8, textColor=dark_gray, alignment=TA_LEFT, leading=11)
    cell_right = ParagraphStyle("tdr", fontName="Helvetica", fontSize=8, textColor=dark_gray, alignment=TA_RIGHT, leading=11)

    table_data = [[Paragraph(h, header_style) for h in sale_headers]]

    for row_idx, item in enumerate(sold_items):
        fees_total = (item.platform_fees or Decimal("0")) + (item.shipping_cost or Decimal("0"))
        row = [
            Paragraph(item.name[:35] + ("…" if len(item.name) > 35 else ""), cell_style),
            Paragraph(item.platform, cell_style),
            Paragraph(_fmt_date(item.sale_date), cell_style),
            Paragraph(_fmt_eur(item.purchase_price), cell_right),
            Paragraph(_fmt_eur(item.sale_price), cell_right),
            Paragraph(_fmt_eur(fees_total), cell_right),
            Paragraph(_fmt_eur(item.net_profit), cell_right),
        ]
        table_data.append(row)

    if not sold_items:
        table_data.append(
            [Paragraph("Aucune vente pour cette année.", cell_style)] + [""] * 6
        )

    sale_table = Table(table_data, colWidths=col_widths_pdf, repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), indigo),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E5E7EB")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
    ]
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            style_cmds.append(("BACKGROUND", (0, i), (-1, i), light_gray))
    sale_table.setStyle(TableStyle(style_cmds))
    story.append(sale_table)
    story.append(Spacer(1, 0.4 * cm))

    # ------------------------------------------------------------------
    # DAC7 seuils
    # ------------------------------------------------------------------
    story.append(Paragraph("Seuils DAC7", style_section))

    bar_width = page_width * 0.65
    bar_height = 12

    def _progress_bar_table(label: str, pct: float, current: str, max_val: str) -> Table:
        fill_pct = min(pct / 100.0, 1.0)
        filled_w = bar_width * fill_pct
        empty_w = bar_width * (1.0 - fill_pct)

        bar_color = green if pct < 70 else (orange if pct < 85 else red)
        status_text, _ = _dac7_status(pct)

        bar_inner_data = [[""]]
        if filled_w > 0:
            filled_tbl = Table([[""]], colWidths=[filled_w], rowHeights=[bar_height])
            filled_tbl.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), bar_color)]))
        if empty_w > 0:
            empty_tbl = Table([[""]], colWidths=[empty_w], rowHeights=[bar_height])
            empty_tbl.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#E5E7EB"))]))

        bar_cells = []
        if filled_w > 0:
            bar_cells.append(filled_tbl)
        if empty_w > 0:
            bar_cells.append(empty_tbl)

        bar_row = bar_cells if bar_cells else [Paragraph("", style_normal)]
        if len(bar_cells) == 2:
            bar_combined = Table([bar_cells], colWidths=[filled_w, empty_w], rowHeights=[bar_height])
        elif len(bar_cells) == 1:
            w = filled_w if filled_w > 0 else empty_w
            bar_combined = Table([bar_cells], colWidths=[w], rowHeights=[bar_height])
        else:
            bar_combined = Table([[""]], colWidths=[bar_width], rowHeights=[bar_height])
        bar_combined.setStyle(TableStyle([("TOPPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 0), ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0)]))

        pct_str = f"{pct:.1f}%"
        row_data = [
            [
                Paragraph(label, ParagraphStyle("lbl", fontName="Helvetica-Bold", fontSize=9, textColor=dark_gray)),
                bar_combined,
                Paragraph(pct_str, ParagraphStyle("pct", fontName="Helvetica-Bold", fontSize=9, textColor=bar_color, alignment=TA_RIGHT)),
                Paragraph(f"{current} / {max_val}", ParagraphStyle("vals", fontName="Helvetica", fontSize=8, textColor=dark_gray)),
            ]
        ]
        t = Table(row_data, colWidths=[page_width * 0.18, bar_width, page_width * 0.08, page_width * 0.09])
        t.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        return t

    story.append(_progress_bar_table(
        "Recettes brutes",
        rx_pct,
        _fmt_eur(gross_receipts),
        _fmt_eur(DAC7_RECEIPTS),
    ))
    story.append(Spacer(1, 3 * mm))
    story.append(_progress_bar_table(
        "Transactions",
        tx_pct,
        str(total_sold),
        str(DAC7_TRANSACTIONS),
    ))
    story.append(Spacer(1, 0.4 * cm))

    # ------------------------------------------------------------------
    # Notes comptables
    # ------------------------------------------------------------------
    story.append(Paragraph("Notes comptables", style_section))
    notes_text = (
        "Les recettes brutes correspondent au montant total des ventes (prix de vente), "
        "conformément à la directive DAC7 (UE). La plateforme de vente est tenue de transmettre "
        "ces informations à la DGFIP (Direction Générale des Finances Publiques) dès qu'un vendeur "
        "atteint 30 transactions <b>ou</b> 2 000 € de recettes brutes sur l'année civile.<br/><br/>"
        "Une fois ce seuil atteint, l'intégralité des revenus (depuis le premier euro) peut être "
        "imposable. Le bénéfice net (recette brute – prix d'achat – frais plateforme – frais de port) "
        "constitue votre base imposable à reporter sur votre déclaration 2042 C PRO ou votre déclaration "
        "spécifique en régime micro-BNC / BIC selon votre situation.<br/><br/>"
        "<i>Ce document est fourni à titre informatif et ne constitue pas un conseil fiscal. "
        "Consultez un expert-comptable pour votre situation personnelle.</i>"
    )
    story.append(Paragraph(notes_text, style_note))

    # ------------------------------------------------------------------
    # Footer function
    # ------------------------------------------------------------------
    footer_text = f"Document généré par Revendu - revendu.fr  |  Confidentiel  |  {user.full_name}  |  {year}"

    def _add_footer(canvas, doc):  # noqa: ANN001
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#9CA3AF"))
        canvas.drawCentredString(A4[0] / 2, 1.2 * cm, footer_text)
        canvas.setLineWidth(0.3)
        canvas.setStrokeColor(colors.HexColor("#E5E7EB"))
        canvas.line(2 * cm, 1.6 * cm, A4[0] - 2 * cm, 1.6 * cm)
        # Page number
        page_num = canvas.getPageNumber()
        canvas.drawRightString(A4[0] - 2 * cm, 1.2 * cm, f"Page {page_num}")
        canvas.restoreState()

    doc.build(story, onFirstPage=_add_footer, onLaterPages=_add_footer)
    buf.seek(0)
    return buf.read()
